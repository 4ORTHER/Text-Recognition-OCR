import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Sample DataFrame
data = {
    'Column1': [1, 2, 3],
    'Column2': ['A', 'B', 'C']
}

df = pd.DataFrame(data)

# Create a Pandas Excel writer using XlsxWriter as the engine
excel_writer = pd.ExcelWriter('output.xlsx', engine='openpyxl')

# Convert the DataFrame to an XlsxWriter Excel object
df.to_excel(excel_writer, index=True, index_label='Index', sheet_name='Sheet1')

# Get the xlsxwriter workbook and worksheet objects
workbook = excel_writer.book
worksheet = excel_writer.sheets['Sheet1']

# Define colors in aRGB hex format
colors = ['FFFF5733', 'FF33FF57', 'FF3366FF']

# Apply colors to the cells in the 'Column1' column
for idx, color in enumerate(colors, start=1):
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    column_index = df.columns.get_loc('Column1') + 2  # Excel columns are 1-based, so add 2
    for row_idx, value in enumerate(df['Column1'], start=2):  # Start from row 2 (Excel row 2)
        if value == idx:
            worksheet.cell(row=row_idx, column=column_index).fill = fill

# Save the workbook to the file
workbook.save('output.xlsx')

print("Excel file with colored cells has been created.")
